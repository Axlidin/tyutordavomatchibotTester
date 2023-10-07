from asyncio import sleep

import openpyxl
from aiogram.types import Message

from data.config import ADMINS
from loader import dp, db, bot


@dp.message_handler(text="📥 Yuklash|Download")
async def Download_db_to_excel(message: Message):
    tg_id = message.from_user.id
    if tg_id == ADMINS[0]:
        book = openpyxl.Workbook()
        sheet = book.active
        # Ustun nomlarini o'zgaruvchi sifatida e'lon qilamiz
        headers = ["№", "Fakultet", "FIO",
                   "Davomat", "Vaqt", "Tekshirdi", "Telegram id"]
        # Ustun nomlarini yozish
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
        file = await db.select_all_davomat()
        # print(file)
        i = 1
        for row in file:
            i += 1
            j = 1
            for col in row:
                cell = sheet.cell(row=i, column=j)
                cell.value = col
                j += 1

        # (D) SAVE EXCEL FILE & CLOSE DB
        book.save("adu_tyutorlar_davomat.xlsx")
        book.close()
        msg = await message.answer("fayl yuklanmoqda....")
        await sleep(5)
        await msg.delete()
        await bot.send_document(chat_id=message.from_user.id, document=open("adu_tyutorlar_davomat.xlsx", 'rb'))
    elif tg_id == ADMINS[1]:
        book = openpyxl.Workbook()
        sheet = book.active
        # Ustun nomlarini o'zgaruvchi sifatida e'lon qilamiz
        headers = ["№", "Fakultet", "FIO",
                   "Davomat", "Vaqt", "Tekshirdi", "Telegram id"]
        # Ustun nomlarini yozish
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
        file = await db.select_all_davomat()
        i = 1
        for row in file:
            i += 1
            j = 1
            for col in row:
                cell = sheet.cell(row=i, column=j)
                cell.value = col
                j += 1

        # (D) SAVE EXCEL FILE & CLOSE DB
        book.save("adu_tyutorlar_davomat.xlsx")
        book.close()
        msg = await message.answer("fayl yuklanmoqda....")
        await sleep(5)
        await msg.delete()
        await bot.send_document(chat_id=message.from_user.id, document=open("adu_tyutorlar_davomat.xlsx", 'rb'))
    elif tg_id == 5287995033:
        book = openpyxl.Workbook()
        sheet = book.active
        # Ustun nomlarini o'zgaruvchi sifatida e'lon qilamiz
        headers = ["№", "Fakultet", "FIO",
                   "Davomat", "Vaqt", "Tekshirdi", "Telegram id"]
        # Ustun nomlarini yozish
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
        faculty = 'atki'
        file = await db.admin_TYUTOR(faculty=faculty)
        i = 1
        for row in file:
            i += 1
            j = 1
            for col in row:
                cell = sheet.cell(row=i, column=j)
                cell.value = col
                j += 1

        # (D) SAVE EXCEL FILE & CLOSE DB
        book.save("atki_tyutorlar_davomat.xlsx")
        book.close()
        msg = await message.answer("fayl yuklanmoqda....")
        await sleep(5)
        await msg.delete()
        await bot.send_document(chat_id=message.from_user.id, document=open("atki_tyutorlar_davomat.xlsx", 'rb'))
    elif tg_id == 433569893:
        book = openpyxl.Workbook()
        sheet = book.active
        # Ustun nomlarini o'zgaruvchi sifatida e'lon qilamiz
        headers = ["№", "Fakultet", "FIO",
                   "Davomat", "Vaqt", "Tekshirdi", "Telegram id"]
        # Ustun nomlarini yozish
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
        faculty = 'tarix'
        file = await db.admin_TYUTOR(faculty=faculty)
        i = 1
        for row in file:
            i += 1
            j = 1
            for col in row:
                cell = sheet.cell(row=i, column=j)
                cell.value = col
                j += 1

        # (D) SAVE EXCEL FILE & CLOSE DB
        book.save("tarix_tyutorlar_davomat.xlsx")
        book.close()
        msg = await message.answer("fayl yuklanmoqda....")
        await sleep(5)
        await msg.delete()
        await bot.send_document(chat_id=message.from_user.id, document=open("tarix_tyutorlar_davomat.xlsx", 'rb'))
    elif tg_id == 313884048:
        book = openpyxl.Workbook()
        sheet = book.active
        # Ustun nomlarini o'zgaruvchi sifatida e'lon qilamiz
        headers = ["№", "Fakultet", "FIO",
                   "Davomat", "Vaqt", "Tekshirdi", "Telegram id"]
        # Ustun nomlarini yozish
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
        faculty = 'tabiiy fanlar'
        file = await db.admin_TYUTOR(faculty=faculty)
        i = 1
        for row in file:
            i += 1
            j = 1
            for col in row:
                cell = sheet.cell(row=i, column=j)
                cell.value = col
                j += 1

        # (D) SAVE EXCEL FILE & CLOSE DB
        book.save("tabiiy_fanlar_tyutorlar_davomat.xlsx")
        book.close()
        msg = await message.answer("fayl yuklanmoqda....")
        await sleep(5)
        await msg.delete()
        await bot.send_document(chat_id=message.from_user.id, document=open("tabiiy_fanlar_tyutorlar_davomat.xlsx", 'rb'))
    elif tg_id == 591486700:
        book = openpyxl.Workbook()
        sheet = book.active
        # Ustun nomlarini o'zgaruvchi sifatida e'lon qilamiz
        headers = ["№", "Fakultet", "FIO",
                   "Davomat", "Vaqt", "Tekshirdi", "Telegram id"]
        # Ustun nomlarini yozish
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
        faculty = 'fizika-matematika'
        file = await db.admin_TYUTOR(faculty=faculty)
        i = 1
        for row in file:
            i += 1
            j = 1
            for col in row:
                cell = sheet.cell(row=i, column=j)
                cell.value = col
                j += 1

        # (D) SAVE EXCEL FILE & CLOSE DB
        book.save("fizika_matematika_tyutorlar_davomat.xlsx")
        book.close()
        msg = await message.answer("fayl yuklanmoqda....")
        await sleep(5)
        await msg.delete()
        await bot.send_document(chat_id=message.from_user.id, document=open("fizika_matematika_tyutorlar_davomat.xlsx", 'rb'))
    elif tg_id == 564495980:
        book = openpyxl.Workbook()
        sheet = book.active
        # Ustun nomlarini o'zgaruvchi sifatida e'lon qilamiz
        headers = ["№", "Fakultet", "FIO",
                   "Davomat", "Vaqt", "Tekshirdi", "Telegram id"]
        # Ustun nomlarini yozish
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
        faculty = 'filologiya'
        file = await db.admin_TYUTOR(faculty=faculty)
        i = 1
        for row in file:
            i += 1
            j = 1
            for col in row:
                cell = sheet.cell(row=i, column=j)
                cell.value = col
                j += 1

        # (D) SAVE EXCEL FILE & CLOSE DB
        book.save("filologiya_tyutorlar_davomat.xlsx")
        book.close()
        msg = await message.answer("fayl yuklanmoqda....")
        await sleep(5)
        await msg.delete()
        await bot.send_document(chat_id=message.from_user.id, document=open("filologiya_tyutorlar_davomat.xlsx", 'rb'))
    elif tg_id == 398900136:
        book = openpyxl.Workbook()
        sheet = book.active
        # Ustun nomlarini o'zgaruvchi sifatida e'lon qilamiz
        headers = ["№", "Fakultet", "FIO",
                   "Davomat", "Vaqt", "Tekshirdi", "Telegram id"]
        # Ustun nomlarini yozish
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
        faculty = "jismoniy ma'daniyat"
        file = await db.admin_TYUTOR(faculty=faculty)
        i = 1
        for row in file:
            i += 1
            j = 1
            for col in row:
                cell = sheet.cell(row=i, column=j)
                cell.value = col
                j += 1

        # (D) SAVE EXCEL FILE & CLOSE DB
        book.save("jismoniy_ma_daniyat_tyutorlar_davomat.xlsx")
        book.close()
        msg = await message.answer("fayl yuklanmoqda....")
        await sleep(5)
        await msg.delete()
        await bot.send_document(chat_id=message.from_user.id, document=open("jismoniy_ma_daniyat_tyutorlar_davomat.xlsx", 'rb'))
    elif tg_id == 744216941:
        book = openpyxl.Workbook()
        sheet = book.active
        # Ustun nomlarini o'zgaruvchi sifatida e'lon qilamiz
        headers = ["№", "Fakultet", "FIO",
                   "Davomat", "Vaqt", "Tekshirdi", "Telegram id"]
        # Ustun nomlarini yozish
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
        faculty = "ijtimoiy iqtisodiyot"
        file = await db.admin_TYUTOR(faculty=faculty)
        i = 1
        for row in file:
            i += 1
            j = 1
            for col in row:
                cell = sheet.cell(row=i, column=j)
                cell.value = col
                j += 1

        # (D) SAVE EXCEL FILE & CLOSE DB
        book.save("ijtimoiy_iqtisodiyot_tyutorlar_davomat.xlsx")
        book.close()
        msg = await message.answer("fayl yuklanmoqda....")
        await sleep(5)
        await msg.delete()
        await bot.send_document(chat_id=message.from_user.id, document=open("ijtimoiy_iqtisodiyot_tyutorlar_davomat.xlsx", 'rb'))
    elif tg_id == 187805821:
        book = openpyxl.Workbook()
        sheet = book.active
        # Ustun nomlarini o'zgaruvchi sifatida e'lon qilamiz
        headers = ["№", "Fakultet", "FIO",
                   "Davomat", "Vaqt", "Tekshirdi", "Telegram id"]
        # Ustun nomlarini yozish
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
        faculty = "san'atshunoslik"
        file = await db.admin_TYUTOR(faculty=faculty)
        i = 1
        for row in file:
            i += 1
            j = 1
            for col in row:
                cell = sheet.cell(row=i, column=j)
                cell.value = col
                j += 1

        # (D) SAVE EXCEL FILE & CLOSE DB
        book.save("san_atshunoslik_tyutorlar_davomat.xlsx")
        book.close()
        msg = await message.answer("fayl yuklanmoqda....")
        await sleep(5)
        await msg.delete()
        await bot.send_document(chat_id=message.from_user.id, document=open("san_atshunoslik_tyutorlar_davomat.xlsx", 'rb'))
    elif tg_id == 6029202056:
        book = openpyxl.Workbook()
        sheet = book.active
        # Ustun nomlarini o'zgaruvchi sifatida e'lon qilamiz
        headers = ["№", "Fakultet", "FIO",
                   "Davomat", "Vaqt", "Tekshirdi", "Telegram id"]
        # Ustun nomlarini yozish
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
        faculty = "pedagogika"
        file = await db.admin_TYUTOR(faculty=faculty)
        i = 1
        for row in file:
            i += 1
            j = 1
            for col in row:
                cell = sheet.cell(row=i, column=j)
                cell.value = col
                j += 1

        # (D) SAVE EXCEL FILE & CLOSE DB
        book.save("pedagogika_tyutorlar_davomat.xlsx")
        book.close()
        msg = await message.answer("fayl yuklanmoqda....")
        await sleep(5)
        await msg.delete()
        await bot.send_document(chat_id=message.from_user.id, document=open("pedagogika_tyutorlar_davomat.xlsx", 'rb'))
    else:
        pass